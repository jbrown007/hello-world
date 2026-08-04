"""Regenerate the Excel watchlist from data/*.yaml.

Data lives in YAML; this module is presentation only. Edit the YAML,
rerun `ff build`, and the workbook is rebuilt from scratch.
"""
from __future__ import annotations
import pathlib
import yaml
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from .config import DATA, ROOT, league, byes, as_range, is_confirmed

FONT = "Arial"
NAVY = PatternFill("solid", fgColor="1F3864")
TIER_FILL = {
    "Tier 1": PatternFill("solid", fgColor="F4CCCC"),
    "Tier 2": PatternFill("solid", fgColor="FCE5CD"),
    "Tier 3": PatternFill("solid", fgColor="D9EAD3"),
}
THIN = Border(*[Side("thin", color="BFBFBF")] * 4)
WRAP = Alignment(wrap_text=True, vertical="top")

COLUMNS = [
    ("priority", "Priority", 10), ("pos", "Pos", 6), ("team", "Team", 16),
    ("bye", "Bye", 6), ("situation", "Situation", 34), ("players", "Players in Play", 40),
    ("why", "Why It Matters", 40), ("read", "Current Read", 36),
    ("watch_for", "Watch For", 34), ("implication", "Draft Implication", 38),
    ("confidence", "Confidence", 11), ("status", "Status", 11), ("notes", "Notes", 60),
]


def _load(name):
    with (DATA / f"{name}.yaml").open(encoding="utf-8") as fh:
        return yaml.safe_load(fh)


def build(out: pathlib.Path | None = None) -> pathlib.Path:
    out = pathlib.Path(out) if out else ROOT / "build" / "2026_Camp_Watchlist.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)

    rows = _load("watchlist")
    lg = league()
    wb = Workbook()

    _start_here(wb.active, lg)
    _watchlist(wb.create_sheet("Watchlist"), rows)
    _rb_board(wb.create_sheet("RB Board"), _load("rb_board"))
    _wr_board(wb.create_sheet("WR Board"), _load("wr_board"))
    _te_board(wb.create_sheet("TE Board"), _load("te_board"))
    _screen(wb.create_sheet("Screen"), _load("screen"))

    wb.save(out)
    return out


def _cell(ws, r, c, v, bold=False, size=10, fill=None, color=None):
    x = ws.cell(r, c, v)
    x.font = Font(name=FONT, bold=bold, size=size, color=color)
    x.alignment = WRAP
    if fill:
        x.fill = fill
    return x


def _start_here(ws, lg):
    ws.title = "Start Here"
    s, p = lg["season"], lg["payouts"]
    _cell(ws, 2, 2, "2026 CAMP WATCHLIST", bold=True, size=14)
    _cell(ws, 3, 2, f"{lg['teams']}-team | {lg['scoring']} | superflex | {lg['roster_size']}-man | IR slots: {lg['ir_slots']}")
    def _fmt(v):
        r = as_range(v)
        return str(r[0]) if len(r) == 1 else "/".join(map(str, r)) + " (UNCONFIRMED)"

    _cell(ws, 4, 2, f"Regular season W1-{_fmt(s['regular_weeks'])} | Playoffs W{_fmt(s['playoff_start'])}-{s['playoff_end']} | Weekly payouts W1-{p['weekly_payout_weeks']}")

    _cell(ws, 6, 2, "DASHBOARD", bold=True, size=12)
    for i, t in enumerate(["Tier 1 (QB)", "Tier 2 (RB)", "Tier 3 (WR/TE)", "Total"]):
        _cell(ws, 7, 3 + i, t, bold=True, color="FFFFFF").fill = NAVY
    # Status column carries three values, not two. Counting only Unsettled and
    # Resolved silently hid every Trending row from the dashboard.
    labels = [("Situations tracked", None), ("Still unsettled", "Unsettled"),
              ("Trending", "Trending"), ("Resolved", "Resolved")]
    for r, (label, status) in enumerate(labels, start=8):
        _cell(ws, r, 2, label, bold=True)
        for i, tier in enumerate(["Tier 1", "Tier 2", "Tier 3"]):
            f = (f'=COUNTIF(Watchlist!$A$2:$A$200,"{tier}*")' if not status
                 else f'=COUNTIFS(Watchlist!$A$2:$A$200,"{tier}*",Watchlist!$L$2:$L$200,"{status}")')
            _cell(ws, r, 3 + i, f)
        _cell(ws, r, 6, '=COUNTA(Watchlist!$E$2:$E$200)' if not status
              else f'=COUNTIF(Watchlist!$L$2:$L$200,"{status}")')

    _cell(ws, 12, 2, "BYE LANDMINES", bold=True, size=12)
    bad = max(byes().items(), key=lambda kv: len(kv[1]))
    _cell(ws, 13, 2, f"Week {bad[0]}")
    _cell(ws, 13, 3, f"{len(bad[1])} teams out: {', '.join(bad[1])}. Cap yourself at two starters from this group.")
    seed = "; ".join(
        f"W{w}: {', '.join(byes().get(w, [])) or 'no byes - clean'}" for w in as_range(s["regular_weeks"])
    )
    _cell(ws, 14, 2, "Seeding week")
    _cell(ws, 14, 3, ("" if is_confirmed(s["regular_weeks"]) else "UNCONFIRMED. ") +
          f"Last regular-season week decides the top-two bye. {seed}")
    pw = sorted({w for st in as_range(s["playoff_start"]) for w in byes() if st <= w <= s["playoff_end"]})
    _cell(ws, 15, 2, "Playoff window")
    _cell(ws, 15, 3, ("" if is_confirmed(s["playoff_start"]) else "UNCONFIRMED. ") +
          ("; ".join(f"W{w}: {', '.join(byes()[w])}" for w in pw) or "No bye conflicts in the playoff window."))

    for col, width in zip("BCDEF", [24, 30, 22, 22, 22]):
        ws.column_dimensions[col].width = width


def _watchlist(ws, rows):
    for c, (_, header, width) in enumerate(COLUMNS, start=1):
        x = _cell(ws, 1, c, header, bold=True, color="FFFFFF")
        x.fill = NAVY
        x.border = THIN
        ws.column_dimensions[x.column_letter].width = width
    for r, row in enumerate(rows, start=2):
        fill = TIER_FILL.get(str(row.get("priority", ""))[:6])
        for c, (key, _, _) in enumerate(COLUMNS, start=1):
            x = _cell(ws, r, c, row.get(key, ""), fill=fill)
            x.border = THIN
        ws.row_dimensions[r].height = 46
    ws.freeze_panes = "A2"


def _stack_cap(ws, r, board):
    # Caps live in data/stack_caps.yaml (single source); boards carry a pointer.
    _cell(ws, r, 2, "STACK CAP", bold=True, size=12)
    _caps = _load("stack_caps")
    _named = _caps["named"] if isinstance(_caps, dict) else _caps
    _gen = _caps.get("general", {}) if isinstance(_caps, dict) else {}
    for cap in _named:
        _cell(ws, r + 1, 2, f"{cap['team']} - bye W{cap['bye']}, max {cap['max_starters']} starters")
        _cell(ws, r + 1, 3, f"{cap['why']} {cap.get('resolved_note', '')}", fill=TIER_FILL["Tier 1"])
        r += 1
    if _gen:
        _cell(ws, r + 1, 2, f"ANY team: {_gen.get('flag_at', 3)}+ players = FLAG")
        _cell(ws, r + 1, 3, " ".join(str(_gen.get("why", "")).split()))
        r += 1
    _cell(ws, r + 1, 3, board.get("stack_cap_ref", ""))
    return r + 3


def _rb_board(ws, board):
    _cell(ws, 2, 2, "RB BOARD", bold=True, size=14)
    _cell(ws, 3, 2, "PROVENANCE", bold=True)
    _cell(ws, 3, 3, board["provenance"], fill=TIER_FILL["Tier 1"])
    _cell(ws, 4, 2, "Rule", bold=True)
    _cell(ws, 4, 3, board["rule"]["note"])
    _cell(ws, 5, 2, f"Window R{board['window']['rounds']}", bold=True)
    _cell(ws, 5, 3, board["window"]["note"])

    r = 7
    _cell(ws, r, 2, "TIERS", bold=True, size=12)
    r += 1
    for t in board["tiers"]:
        _cell(ws, r, 2, f"Tier {t['tier']}", bold=True)
        _cell(ws, r, 3, f"{', '.join(t['who']) or '(empty)'} - {t['note']}")
        r += 1

    for title, key, fill in (("TARGETS", "targets", TIER_FILL["Tier 3"]),
                             ("FADES", "fades", TIER_FILL["Tier 1"])):
        r += 1
        _cell(ws, r, 2, title, bold=True, size=12)
        r += 1
        headers = ["Player", "Team", "ADP", "Round", "Why"]
        for c, h in enumerate(headers, start=2):
            x = _cell(ws, r, c, h, bold=True, color="FFFFFF")
            x.fill = NAVY
            x.border = THIN
        r += 1
        for p in board[key]:
            vals = [p["player"], p["team"], p.get("adp"), p.get("round"), p["why"]]
            for c, v in enumerate(vals, start=2):
                x = _cell(ws, r, c, "" if v is None else v, fill=fill)
                x.border = THIN
            r += 1

    r += 1
    _cell(ws, r, 2, "ROUND PLAN", bold=True, size=12)
    r += 1
    for rounds, plan in board["round_plan"].items():
        _cell(ws, r, 2, f"R{rounds}", bold=True)
        _cell(ws, r, 3, plan)
        r += 1

    for col, width in zip("BCDEF", [30, 60, 12, 8, 90]):
        ws.column_dimensions[col].width = width


def _wr_board(ws, board):
    _cell(ws, 2, 2, "WR BOARD", bold=True, size=14)
    _cell(ws, 3, 2, f"Cliff at {board['cliff']['at']}", bold=True)
    _cell(ws, 3, 3, board["cliff"]["note"])
    _cell(ws, 4, 2, f"{board['rule']['targets_per_game']} targets/game", bold=True)
    _cell(ws, 4, 3, board["rule"]["note"])
    _cell(ws, 5, 2, "Context", bold=True)
    _cell(ws, 5, 3, board["context"])

    _cell(ws, 7, 2, "VALUE BOARD", bold=True, size=12)
    headers = ["Player", "Team", "Target Rank", "ADP", "Gap", "Round", "Flag / Note"]
    for c, h in enumerate(headers, start=2):
        x = _cell(ws, 8, c, h, bold=True, color="FFFFFF")
        x.fill = NAVY
        x.border = THIN
    r = 9
    for p in board["value_board"]:
        flag = "; ".join(str(p[k]) for k in ("flag", "note") if p.get(k))
        vals = [p["player"], p["team"], p.get("target_rank"), p.get("adp"),
                p.get("gap"), p.get("round"), flag]
        for c, v in enumerate(vals, start=2):
            x = _cell(ws, r, c, "" if v is None else v)
            x.border = THIN
        r += 1

    r = _stack_cap(ws, r + 1, board)

    _cell(ws, r, 2, "ROUND PLAN", bold=True, size=12)
    r += 1
    for rounds, plan in board["round_plan"].items():
        _cell(ws, r, 2, f"R{rounds}", bold=True)
        _cell(ws, r, 3, plan)
        r += 1

    for col, width in zip("BCDEFGH", [22, 90, 12, 8, 6, 7, 60]):
        ws.column_dimensions[col].width = width


def _te_board(ws, board):
    _cell(ws, 2, 2, "TE BOARD", bold=True, size=14)
    _cell(ws, 3, 2, f"Cliff at {board['cliff']['at']}", bold=True)
    _cell(ws, 3, 3, board["cliff"]["note"])
    _cell(ws, 4, 2, "Qualifying test", bold=True)
    _cell(ws, 4, 3, board["qualifying_test"])
    _cell(ws, 5, 2, "Barbell flaw", bold=True)
    _cell(ws, 5, 3, board["barbell_flaw"])

    _cell(ws, 7, 2, "PATHS", bold=True, size=12)
    headers = ["Path", "Name", "Cost", "Who", "Verdict"]
    for c, h in enumerate(headers, start=2):
        x = _cell(ws, 8, c, h, bold=True, color="FFFFFF")
        x.fill = NAVY
        x.border = THIN
    r = 9
    for p in board["paths"]:
        fill = TIER_FILL["Tier 3"] if "RECOMMENDED" in p["verdict"] else None
        vals = [p["id"], p["name"], p["cost"], ", ".join(p["who"]), p["verdict"]]
        for c, v in enumerate(vals, start=2):
            x = _cell(ws, r, c, v, fill=fill)
            x.border = THIN
        r += 1

    r = _stack_cap(ws, r + 1, board)

    _cell(ws, r, 2, "THE CALL", bold=True, size=12)
    _cell(ws, r, 3, board["call"], fill=TIER_FILL["Tier 3"])

    for col, width in zip("BCDEF", [14, 90, 14, 36, 50]):
        ws.column_dimensions[col].width = width


def _screen(ws, teams):
    headers = ["Team", "New HC", "New OC", "T1", "T2", "T3", "T4", "T5", "T6", "Hits", "Priority", "Note"]
    for c, h in enumerate(headers, start=1):
        x = _cell(ws, 1, c, h, bold=True, color="FFFFFF")
        x.fill = NAVY
    for r, t in enumerate(teams, start=2):
        trig = [t.get(f"t{i}", "") for i in range(1, 7)]
        hits = sum(1 for v in trig if str(v).upper() == "Y")
        pri = "HIGH" if hits >= 3 else ("MED" if hits == 2 else "LOW")
        vals = [t.get("team"), t.get("new_hc", ""), t.get("new_oc", "")] + trig + [hits, pri, t.get("note", "")]
        for c, v in enumerate(vals, start=1):
            x = _cell(ws, r, c, v)
            if c == 11:
                x.fill = {"HIGH": TIER_FILL["Tier 1"], "MED": TIER_FILL["Tier 2"]}.get(pri, TIER_FILL["Tier 3"])
    for col, width in zip("ABCDEFGHIJKL", [16, 16, 20, 5, 5, 5, 5, 5, 5, 7, 10, 70]):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
