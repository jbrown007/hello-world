#!/usr/bin/env python3
"""Automated integrity checks for the ff2026 framework.

Run first, before any manual review:

    cd ff2026 && python3 tests/verify.py

Exit code 0 means every structural check passed. It does NOT mean the football
analysis is correct - see HANDOFF.md for what still needs human eyes.
"""
from __future__ import annotations

import io
import subprocess
import sys
import traceback
from contextlib import redirect_stdout
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

NFL_TEAMS = {
    "ARI", "ATL", "BAL", "BUF", "CAR", "CHI", "CIN", "CLE", "DAL", "DEN", "DET",
    "GB", "HOU", "IND", "JAX", "KC", "LAC", "LAR", "LV", "MIA", "MIN", "NE",
    "NO", "NYG", "NYJ", "PHI", "PIT", "SEA", "SF", "TB", "TEN", "WSH",
}

results: list[tuple[str, bool, str]] = []


def check(name):
    """Decorator that records pass/fail instead of raising."""
    def wrap(fn):
        try:
            detail = fn() or ""
            results.append((name, True, str(detail)))
        except AssertionError as e:
            results.append((name, False, str(e)))
        except Exception:
            results.append((name, False, traceback.format_exc(limit=2).strip().splitlines()[-1]))
        return fn
    return wrap


# --------------------------------------------------------------- data layer
@check("all YAML files parse")
def _():
    import yaml
    files = sorted((ROOT / "data").glob("*.yaml"))
    assert files, "no YAML files found in data/"
    for f in files:
        with f.open(encoding="utf-8") as fh:
            yaml.safe_load(fh)
    return f"{len(files)} files"


@check("league.yaml has every required key")
def _():
    from ffcli.config import league
    lg = league()
    for key in ("teams", "scoring", "superflex", "roster_size", "ir_slots",
                "starters", "draft", "waivers", "season", "payouts"):
        assert key in lg, f"missing top-level key: {key}"
    for key in ("regular_weeks", "playoff_start", "playoff_end"):
        assert key in lg["season"], f"missing season.{key}"
    return f"{lg['teams']} teams, superflex={lg['superflex']}, ir_slots={lg['ir_slots']}"


@check("byes.yaml covers all 32 NFL teams exactly once")
def _():
    from ffcli.config import byes
    seen: list[str] = []
    for teams in byes().values():
        seen += teams
    dupes = {t for t in seen if seen.count(t) > 1}
    assert not dupes, f"team on multiple byes: {sorted(dupes)}"
    missing = NFL_TEAMS - set(seen)
    unknown = set(seen) - NFL_TEAMS
    assert not missing, f"teams with no bye: {sorted(missing)}"
    assert not unknown, f"unrecognised abbreviations: {sorted(unknown)}"
    return f"{len(seen)} teams across {len(byes())} weeks"


@check("bye weeks fall in a plausible range")
def _():
    from ffcli.config import byes
    weeks = sorted(byes())
    assert min(weeks) >= 4, f"bye earlier than W4: {min(weeks)}"
    assert max(weeks) <= 14, f"bye later than W14: {max(weeks)}"
    return f"W{min(weeks)}-W{max(weeks)}, no byes in {sorted(set(range(min(weeks), max(weeks)+1)) - set(weeks))}"


@check("season settings are internally consistent")
def _():
    from ffcli.config import league, as_range
    s = league()["season"]
    reg = as_range(s["regular_weeks"])
    start = as_range(s["playoff_start"])
    end = as_range(s["playoff_end"])[0]
    for r in reg:
        for st in start:
            assert st > r, f"playoffs start W{st} but regular season runs {r} weeks - overlap"
            assert st <= end, f"playoff_start W{st} after playoff_end W{end}"
    return f"regular={reg}, playoff_start={start}, end=W{end}"


@check("watchlist rows have required fields")
def _():
    from ffcli.workbook import _load
    rows = _load("watchlist")
    assert rows, "watchlist is empty"
    required = ("priority", "pos", "team", "situation", "status", "confidence")
    for i, row in enumerate(rows, start=1):
        for k in required:
            assert k in row, f"row {i} missing '{k}'"
        assert str(row["status"]) in {"Resolved", "Trending", "Unsettled"}, \
            f"row {i} has unexpected status: {row['status']!r}"
    return f"{len(rows)} rows"


@check("screen covers exactly 32 teams, no duplicates")
def _():
    from ffcli.workbook import _load
    teams = _load("screen")
    names = [t["team"] for t in teams]
    assert len(names) == 32, f"expected 32 teams, found {len(names)}"
    dupes = {n for n in names if names.count(n) > 1}
    assert not dupes, f"duplicate teams: {sorted(dupes)}"
    return "32 teams"


@check("qb_rule thresholds are monotonic and ordered")
def _():
    from ffcli.config import load
    rule = load("qb_rule")
    for band in rule["rules"]:
        gone = [t["gone"] for t in band["thresholds"]]
        assert gone == sorted(gone, reverse=True), \
            f"round<={band['max_round']} thresholds not descending: {gone}"
        assert gone[-1] == 0, f"round<={band['max_round']} has no catch-all 0 threshold"
    rounds = [b["max_round"] for b in rule["rules"]]
    assert rounds == sorted(rounds), f"rule bands out of order: {rounds}"
    assert rule["hard_floor_round"] >= max(rounds), "hard floor is before the last rule band"
    return f"{len(rule['rules'])} bands, floor R{rule['hard_floor_round']}"


@check("draft trees cover every slot exactly once")
def _():
    from ffcli.config import load, league
    n = league()["teams"]
    covered: list[int] = []
    for branch in load("trees"):
        covered += branch["slots"]
    dupes = {s for s in covered if covered.count(s) > 1}
    assert not dupes, f"slots in multiple branches: {sorted(dupes)}"
    missing = set(range(1, n + 1)) - set(covered)
    assert not missing, f"slots with no branch: {sorted(missing)}"
    return f"slots 1-{n} across {len(load('trees'))} branches"


# --------------------------------------------------------------- logic layer
@check("QB count rule returns a verdict at every boundary")
def _():
    from ffcli.draft import qb_verdict
    from ffcli.config import load
    floor = load("qb_rule")["hard_floor_round"]
    for rnd in range(1, floor + 3):
        for gone in (0, 9, 10, 13, 14, 16, 24):
            v = qb_verdict(rnd, gone)
            assert v.action, f"no action for round={rnd} gone={gone}"
    past = qb_verdict(floor + 1, 0)
    assert past.action == "PAST_FLOOR", f"expected PAST_FLOOR past R{floor}, got {past.action}"
    return f"tested R1-R{floor+2}"


@check("QB rule escalates as more QBs leave the board")
def _():
    from ffcli.draft import qb_verdict
    urgency = {"WAIT": 0, "WAIT_TO_6": 0, "NORMAL": 1, "BY_ROUND_6": 2, "TAKE_QB2_NOW": 3}
    for rnd in (4, 5):
        seen = [urgency[qb_verdict(rnd, g).action] for g in (0, 11, 13, 15, 17, 20)]
        assert seen == sorted(seen), f"round {rnd} urgency not monotonic: {seen}"
    return "monotonic at R4 and R5"


@check("bye audit flags a known-bad roster")
def _():
    from ffcli.byecheck import audit
    res = audit(["IND", "NYJ", "LV", "BAL"], max_per_week=2)
    assert res["warnings"], "four teams on the same bye produced no warning"
    assert 13 in res["grouped"], "W13 grouping missing"
    assert len(res["grouped"][13]) == 4, "expected 4 teams grouped in W13"
    return f"{len(res['warnings'])} warnings, {res['scenarios']} scenarios"


@check("bye audit stays quiet on a clean roster")
def _():
    from ffcli.byecheck import audit
    res = audit(["KC", "CIN", "BUF"], max_per_week=2)
    stack = [w for w in res["warnings"] if "Cap is" in w]
    assert not stack, f"false stacking warning: {stack}"
    return "no false positives"


@check("bye audit rejects nothing silently")
def _():
    from ffcli.byecheck import audit
    res = audit(["IND", "NOTATEAM"], max_per_week=2)
    assert "NOTATEAM" in res["unknown"], "unrecognised team was swallowed"
    return "unknown teams surfaced"


@check("every draft slot returns a branch")
def _():
    from ffcli.draft import tree
    from ffcli.config import league
    for slot in range(1, league()["teams"] + 1):
        b = tree(slot)
        assert b["steps"], f"slot {slot} branch has no steps"
    return "all slots resolve"


@check("weekly sessions 1-4 all render")
def _():
    from ffcli.weekly import session
    for n in (1, 2, 3, 4):
        text = session(n)
        assert "Bring:" in text and "paste this" in text, f"session {n} template malformed"
    return "4 sessions"


# --------------------------------------------------------------- build layer
@check("workbook builds and contains expected tabs")
def _():
    import openpyxl
    from ffcli.workbook import build
    out = build(ROOT / "build" / "_verify.xlsx")
    wb = openpyxl.load_workbook(out)
    for tab in ("Start Here", "Watchlist", "RB Board", "WR Board", "TE Board", "Screen"):
        assert tab in wb.sheetnames, f"missing tab: {tab}"
    ws = wb["Watchlist"]
    assert ws.max_row > 1, "watchlist tab has no data rows"
    assert ws.freeze_panes == "A2", "watchlist header not frozen"
    return f"{wb.sheetnames}, {ws.max_row - 1} rows"


@check("workbook formulas are written as formulas")
def _():
    import openpyxl
    wb = openpyxl.load_workbook(ROOT / "build" / "_verify.xlsx")
    sh = wb["Start Here"]
    formulas = [
        sh.cell(r, c).value
        for r in range(1, 30) for c in range(1, 10)
        if isinstance(sh.cell(r, c).value, str) and sh.cell(r, c).value.startswith("=")
    ]
    assert formulas, "no formulas found - values may have been hardcoded"
    bad = [f for f in formulas if "Watchlist!" not in f]
    assert not bad, f"formulas not referencing Watchlist: {bad}"
    return f"{len(formulas)} formulas, all cross-referencing"


@check("board tabs are generated from the board YAML")
def _():
    """Regression: wr_board.yaml and te_board.yaml were orphaned - no code read
    them, so the workbook's board tabs could silently drift from the data."""
    import openpyxl
    from ffcli.config import load
    wb = openpyxl.load_workbook(ROOT / "build" / "_verify.xlsx")

    def text(sheet):
        return " | ".join(str(c.value) for row in wb[sheet].iter_rows() for c in row if c.value is not None)

    rb, wr, te = text("RB Board"), text("WR Board"), text("TE Board")
    for p in load("wr_board")["value_board"]:
        assert p["player"] in wr, f"WR Board tab missing {p['player']}"
    for path in load("te_board")["paths"]:
        assert path["name"] in te, f"TE Board tab missing path: {path['name']}"
    for p in load("rb_board")["targets"] + load("rb_board")["fades"]:
        assert p["player"] in rb, f"RB Board tab missing {p['player']}"
    assert "provenance" in rb.lower(), "RB Board missing provenance warning (synthesized data)"
    for sheet in ("WR Board", "TE Board"):
        from ffcli.draft import named_caps
        for cap in named_caps():
            assert f"max {cap['max_starters']} starters" in text(sheet), f"{sheet} missing stack cap"
    n_rb = len(load("rb_board")["targets"]) + len(load("rb_board")["fades"])
    return f"{n_rb} RBs, {len(load('wr_board')['value_board'])} WRs, {len(load('te_board')['paths'])} TE paths"


@check("CLI commands all exit cleanly")
def _():
    from ffcli.cli import main
    gpath = ROOT / "build" / "_grade_smoke.txt"
    gpath.parent.mkdir(parents=True, exist_ok=True)
    gpath.write_text("1 RB ATL Bijan Robinson\n2 QB CIN Joe Burrow\n", encoding="utf-8")
    cases = [
        ["settings"], ["confirm"], ["qb", "--round", "4", "--gone", "15"],
        ["grade", str(gpath), "--slot", "7", "--oneqb"],
        ["tree", "--slot", "1"], ["tree", "--slot", "12"],
        ["draft", "--round", "6", "--gone", "16", "--slot", "7"],
        ["rb", "--round", "4", "--held", "1"], ["rb", "--round", "8", "--held", "3"],
        ["qb", "--round", "5", "--gone", "8", "--window", "4"],
        ["sheet", "--slot", "1"], ["sheet", "--slot", "7"], ["sheet", "--slot", "12"],
        ["bye", "IND", "NYJ"], ["weekly", "1"], ["weekly", "4"],
    ]
    for argv in cases:
        buf = io.StringIO()
        with redirect_stdout(buf):
            code = main(argv)
        assert code == 0, f"`ff {' '.join(argv)}` exited {code}"
        assert buf.getvalue().strip(), f"`ff {' '.join(argv)}` printed nothing"
    return f"{len(cases)} commands"


@check("package imports with no side effects")
def _():
    out = subprocess.run(
        [sys.executable, "-c", "import ffcli, ffcli.cli, ffcli.draft, ffcli.byecheck, ffcli.workbook, ffcli.weekly"],
        cwd=ROOT, capture_output=True, text=True,
    )
    assert out.returncode == 0, out.stderr.strip().splitlines()[-1] if out.stderr else "import failed"
    assert not out.stdout.strip(), f"import printed to stdout: {out.stdout!r}"
    return "clean import"


# --------------------------------------------------------------- consistency
@check("watchlist teams resolve to real bye weeks")
def _():
    from ffcli.workbook import _load
    from ffcli.config import bye_of
    NAME = {
        "Titans": "TEN", "Buccaneers": "TB", "Saints": "NO", "Colts": "IND", "Bears": "CHI",
        "Jets": "NYJ", "Raiders": "LV", "Cardinals": "ARI", "Commanders": "WSH",
        "Broncos": "DEN", "Patriots": "NE", "49ers": "SF", "Dolphins": "MIA",
        "Browns": "CLE", "Chargers": "LAC", "Vikings": "MIN", "Seahawks": "SEA",
    }
    unmapped = []
    for row in _load("watchlist"):
        team = str(row.get("team", "")).strip()
        ab = NAME.get(team)
        if ab is None:
            unmapped.append(team)
            continue
        assert bye_of(ab), f"{team} ({ab}) has no bye week"
    return f"unmapped (expected for multi-team rows): {sorted(set(unmapped))}" if unmapped else "all mapped"


@check("no watchlist row claims a bye that contradicts byes.yaml")
def _():
    from ffcli.workbook import _load
    from ffcli.config import bye_of
    NAME = {
        "Titans": "TEN", "Buccaneers": "TB", "Saints": "NO", "Colts": "IND", "Bears": "CHI",
        "Jets": "NYJ", "Raiders": "LV", "Cardinals": "ARI", "Commanders": "WSH",
        "Broncos": "DEN", "Patriots": "NE", "49ers": "SF", "Dolphins": "MIA",
        "Browns": "CLE", "Chargers": "LAC", "Vikings": "MIN", "Seahawks": "SEA",
    }
    bad = []
    for row in _load("watchlist"):
        ab = NAME.get(str(row.get("team", "")).strip())
        stated = row.get("bye")
        if ab and isinstance(stated, int) and bye_of(ab) != stated:
            bad.append(f"{row['team']}: row says W{stated}, byes.yaml says W{bye_of(ab)}")
    assert not bad, "; ".join(bad)
    return "watchlist byes agree with byes.yaml"


@check("unconfirmed settings are reported, not silently defaulted")
def _():
    from ffcli.config import unconfirmed, league, as_range
    pending = unconfirmed()
    s = league()["season"]
    for key in ("regular_weeks", "playoff_start"):
        if len(as_range(s[key])) > 1:
            assert key in pending, f"{key} is a range but not reported by unconfirmed()"
    return f"pending: {sorted(pending) or 'none'}"


@check("round-plan commitments are satisfiable in every branch")
def _():
    """Regression: the boards once committed seven picks to six rounds (audit
    2a) and nothing noticed. Commitments are now data; this proves each
    branch's full set fits one-pick-per-round."""
    from ffcli.draft import commitments_for, satisfiable
    from ffcli.config import load
    details = []
    for label in load("commitments")["branches"]:
        items = commitments_for(label)
        for c in items:
            lo, hi = c["window"]
            assert 1 <= lo <= hi <= 17, f"{label}: bad window {c['window']} for {c['pick']}"
        ok, detail = satisfiable(items)
        assert ok, f"{label}: {detail}"
        details.append(f"{label}={len(items)}")
    assert details, "no branches in commitments.yaml"
    return ", ".join(details) + " commitments, all schedulable"


@check("stack caps are centralized, boards only point")
def _():
    """8/1: caps moved to data/stack_caps.yaml (single source). Boards must
    carry a pointer and must NOT carry their own copy - that duplication is
    how audit 2c's drift happened."""
    from ffcli.config import load
    from ffcli.draft import named_caps, general_cap
    caps = named_caps()
    assert caps, "stack_caps.yaml has no named caps"
    for cap in caps:
        for k in ("team", "bye", "max_starters", "severity", "resolved_note"):
            assert k in cap, f"cap {cap.get('team')} missing {k}"
    gen = general_cap()
    assert isinstance(gen.get("flag_at"), int) and gen["flag_at"] >= 2, \
        "general cap needs an integer flag_at of 2 or more"
    assert gen.get("why"), "general cap has no rationale"
    for cap in caps:
        assert cap["max_starters"] < gen["flag_at"], \
            f"named cap {cap['team']} ({cap['max_starters']}) is not tighter than the " \
            f"general flag ({gen['flag_at']}) - the named cap would be pointless"
    for board in ("wr_board", "te_board"):
        b = load(board)
        assert "stack_cap" not in b, f"{board} still carries an inline stack_cap"
        assert "stack_cap_ref" in b, f"{board} missing the stack_caps pointer"
    return f"{len(caps)} named cap(s) + general flag at {gen['flag_at']}, both boards point"


@check("Warren gate guards every round of his window")
def _():
    """8/4: the gate was written for R5 only, so the slot-6 mock took Warren at
    R4 on one RB - legal by the letter - and missed the 2-by-R4 floor. A gate
    that guards one round of a two-round window is not a gate. It must cover
    Warren's whole commitment window and reach both the sheet and pick screen."""
    from ffcli.draft import warren_gate, commitments_for, sheet, sheet_twocol, draft_screen
    from ffcli.config import load
    g = warren_gate()
    win = next(c["window"] for c in commitments_for("MIDDLE") if "Warren" in c["pick"])
    covered = set(g["applies_rounds"])
    assert covered >= set(range(win[0], win[1] + 1)), \
        f"gate covers {sorted(covered)} but Warren's window is R{win[0]}-R{win[1]}"
    floor = {x["by_end_of_round"]: x["min_held"] for x in load("rb_rule")["rb_floor"]}
    assert g["min_rb_held"] >= floor[max(r for r in floor if r <= win[1])], \
        "gate's RB requirement is looser than the RB floor it protects"
    for rnd in g["applies_rounds"]:
        assert "GATE" in draft_screen(1, rnd, 5), f"pick screen R{rnd} lost the Warren gate"
    # R4 additionally needs the WR banked: QB1+RB+RB+WR fill R1-R4 exactly, so
    # Warren at R4 without it does not risk the WR miss, it guarantees it.
    r4 = min(g["applies_rounds"])
    early = [c for c in commitments_for("EARLY") if c["window"][1] <= r4]
    saturated = len(early) >= r4 and any("WR" in c["pick"] for c in early)
    if saturated:
        # R1-R4 has as many commitments as rounds and one of them is the WR, so
        # spending R4 on Warren cannot leave room for it. The flag is not
        # optional here - the arithmetic requires it.
        assert g.get("r4_needs_wr"), (
            f"R1-R{r4} holds {len(early)} commitments for {r4} rounds including a WR, so "
            "Warren at R4 guarantees the WR miss - r4_needs_wr must be set")
        assert "AND a WR" in draft_screen(1, r4, 5), "pick screen R4 lost the WR condition"
        assert "no WR" in sheet_twocol(3), "two-col sheet R4 lost the WR condition"
    for render in (sheet, sheet_twocol):
        text = render(6)
        hits = [l for l in text.splitlines() if "WARREN WAIT" in l.upper() or "WARREN GATE" in l]
        assert len(hits) >= len(g["applies_rounds"]), \
            f"{render.__name__}: gate shown {len(hits)}x, needs {len(g['applies_rounds'])}"
    return f"gate on R{g['applies_rounds']}, min {g['min_rb_held']} RB, on sheet + pick screen"


@check("grade enforces the 17-spot roster ledger")
def _():
    """Commitments police WHEN a pick lands; nothing policed WHAT the roster
    became, so two slot-4 mocks drafted a TE2 - unstartable here, FLEX excludes
    TE and OP is QB2's - and graded clean. The ledger must sum to the roster
    and a second TE must be called out by name."""
    from ffcli.draft import grade, ledger_report
    from ffcli.config import load, league
    want = load("commitments")["ledger"]
    assert sum(want.values()) == league()["roster_size"], \
        f"ledger sums to {sum(want.values())}, roster is {league()['roster_size']}"
    assert want["TE"] >= 1, "ledger must roster at least the starting TE"
    flex = league()["starters"].get("FLEX")
    assert flex, "no FLEX slot in starters"
    good = ([{"round": 1, "pos": p, "team": "KC", "player": f"{p}{i}"}
             for p, n in want.items() for i in range(n)])
    assert "ledger met" in ledger_report(good), "a ledger-perfect roster did not pass"
    # swap a WR for a second TE: must flag BOTH the over and the short
    bad = [dict(p) for p in good]
    next(p for p in bad if p["pos"] == "WR")["pos"] = "TE"
    rep = ledger_report(bad)
    assert "LEDGER OVER TE: 2 vs 1" in rep, f"TE2 not flagged: {rep}"
    assert "LEDGER SHORT WR: 4 vs 5" in rep, f"WR shortfall not flagged: {rep}"
    assert "never score" not in rep, (
        "grade still claims a TE2 cannot score - FLEX accepts a TE (corrected 8/9)")
    assert "RB6" in rep or "WR5" in rep, (
        "TE2 deviation must be framed against the spot it actually costs")
    assert "ROSTER LEDGER" in grade(bad, "EARLY"), "grade does not print the ledger"
    return f"ledger {sum(want.values())} spots, TE2 + WR shortfall both caught"


@check("live pick screen tracks what you hold and arms the run trigger")
def _():
    """8/13 dry run: at R5 the screen called QB1/RB/RB/WR OVERDUE and told you
    to 'skip any already rostered' - i.e. filter its own false alarms under a
    pick clock. It also could not reach the run trigger; that lived only in
    `ff qb`. Both are the tool's whole job on draft day."""
    from ffcli.draft import (draft_screen, outstanding, parse_have, commitments_for,
                             _commit_name)
    assert parse_have("QB=1,RB=2") == ({"QB": 1, "RB": 2}, set())
    assert parse_have("QB=1,warren") == ({"QB": 1}, {"warren"})
    assert parse_have(None) == ({}, set()) and parse_have("") == ({}, set())
    try:
        parse_have("QB")
        assert False, "parse_have accepted a malformed pair"
    except ValueError:
        pass
    # holding the early picks must retire those commitments, in deadline order
    full = commitments_for("MIDDLE")
    have = {"QB": 1, "RB": 2, "WR": 1}
    left = outstanding("MIDDLE", have)
    assert len(left) == len(full) - 4, f"held 4 picks, {len(full) - len(left)} commitments cleared"
    assert not any(c["pick"].startswith("QB1") for c in left), "QB1 still owed while holding a QB"
    assert any(c["pick"] == "QB2" for c in left), "QB2 wrongly cleared by the QB1 holding"
    # a NAMED commitment must not be cleared by unrelated players at the position
    named = [c for c in full if _commit_name(c["pick"])]
    assert named, "no named commitments to test - Warren/Downs missing from the plan"
    many = outstanding("MIDDLE", {"QB": 3, "RB": 6, "WR": 5, "TE": 1})
    for c in named:
        assert any(x["pick"] == c["pick"] for x in many), \
            f"{c['pick']} cleared by position count alone - it names a specific player"
    keys = {_commit_name(c["pick"]) for c in named}
    assert not outstanding("MIDDLE", {"QB": 3, "RB": 6, "WR": 5, "TE": 1, "K": 1, "DST": 1}, keys), \
        "naming every player plus full counts still leaves commitments owed"
    clean = draft_screen(6, 5, 9, None, have)
    assert "OVERDUE" not in clean, "satisfied commitments still shown as OVERDUE"
    assert "HELD" in clean and "4 of 17 picks made" in clean, "no roster summary"
    # the run trigger must be reachable from the pick screen, not just ff qb
    quiet = draft_screen(6, 5, 7, None, have)
    fired = draft_screen(6, 5, 7, 3, have)
    assert "TAKE_QB2_NOW" not in quiet, "trigger fired without a window"
    assert "RUN DETECTED" in fired, "3-in-12 window did not fire the run trigger"
    # a short RB count must surface without a separate ff rb call
    short = draft_screen(6, 8, 14, None, {"QB": 2, "RB": 2, "WR": 2, "TE": 1})
    assert "TAKE_RB_NOW" in short, "RB floor breach not surfaced on the pick screen"
    return "held-tracking, run trigger and RB floor all live on ff draft"


@check("depth board covers R9-R15 and reaches the sheet")
def _():
    """8/9: an audit found 62% of mock skill picks were on NO board - the plan
    said WHEN but not WHO after R8, which is exactly where the remaining errors
    live. Every depth entry needs a real team, a round window and a verdict;
    the depth rounds must not be left blank on the sheet; and anything on the
    fade list must not also be recommended."""
    from ffcli.draft import depth_at, sheet_twocol, draft_screen
    from ffcli.config import load, bye_of
    b = load("depth_board")
    picks = [p for pos in ("rb", "wr", "te") for p in b.get(pos, [])]
    assert picks, "depth board has no players"
    # Section sizes. An 8/16 edit inserted a new top-level key in the MIDDLE of
    # the wr list, which silently reparented three receivers out of it - and
    # every assertion below still passed, because they only ever looked at what
    # was left. A board that quietly shrinks is the failure mode worth naming.
    for pos, floor in (("rb", 4), ("wr", 4), ("te", 1)):
        assert len(b.get(pos, [])) >= floor, \
            f"depth board {pos} section has {len(b.get(pos, []))} entries, expected at least {floor}"
    seen: dict[str, str] = {}
    for pos in ("rb", "wr", "te", "fades", "graduated"):
        for p in b.get(pos, []):
            assert p["player"] not in seen, \
                f"{p['player']} appears in both '{seen[p['player']]}' and '{pos}'"
            seen[p["player"]] = pos
    for p in picks + b.get("fades", []):
        assert bye_of(p["team"]), f"{p['player']}: {p['team']!r} is not a real team code"
        assert p.get("why"), f"{p['player']} has no rationale"
    for p in picks:
        assert p.get("verdict"), f"{p['player']} has no verdict"
        span = __import__("ffcli.draft", fromlist=["x"])._round_span(p.get("rounds"))
        assert span and span[0] >= 8, f"{p['player']} window {p.get('rounds')} is not a depth round"
    faded = {p["player"] for p in b.get("fades", [])}
    assert not (faded & {p["player"] for p in picks}), \
        f"player both recommended and faded: {faded & {p['player'] for p in picks}}"
    # the rounds that used to read 'free - best value' must now carry names
    blank = [r for r in range(9, 16) if not depth_at(r)]
    assert not blank, f"depth rounds still empty: {blank}"
    text = sheet_twocol(3)
    for r in (9, 12, 15):
        row = next(l for l in text.splitlines() if l.startswith(f"R{r} "))
        assert "free - best value" not in row, f"sheet R{r} still blank"
    assert "STRONG BUY" in draft_screen(3, 9, 8), "pick screen lost depth verdicts"
    # W14 is the seeding week - any depth name on it must be flagged, not silent
    for p in picks:
        if bye_of(p["team"]) == 14:
            assert "CAUTION" in p["verdict"], \
                f"{p['player']} is on the W14 seeding-week bye but carries no caution"
    return f"{len(picks)} depth names + {len(faded)} fades, R9-R15 all covered"


@check("mock log aggregates into a pattern report")
def _():
    """Josh is repping all 12 slots before Sept 7. The log has to answer which
    slots remain, which errors repeat, and whether a fix held - not just store
    rows. Every logged error must be a real recurring-error key."""
    from ffcli.draft import mocks_report
    from ffcli.config import load, league
    rows = load("mocks")
    assert rows, "mocks.yaml empty"
    teams = league()["teams"]
    for i, m in enumerate(rows, start=1):
        for k in ("date", "slot", "format", "notes", "errors"):
            assert k in m, f"mock row {i} missing {k}"
        assert 1 <= m["slot"] <= teams, f"row {i}: slot {m['slot']} outside 1-{teams}"
        assert isinstance(m["errors"], list), f"row {i}: errors must be a list"
        if m.get("score") is not None:
            assert 0 <= m["score"] <= m["of"], f"row {i}: score {m['score']}/{m['of']}"
    text = mocks_report()
    for token in ("SLOT COVERAGE", "SCORES", "RECURRING ERRORS"):
        assert token in text, f"mocks report missing {token}"
    done = {m["slot"] for m in rows}
    missing = [s for s in range(1, teams + 1) if s not in done]
    if missing:
        assert "STILL TO DO" in text, "unrepped slots not surfaced"
        assert str(missing[0]) in text.split("STILL TO DO")[1], "missing slot not listed"
    else:
        assert "ALL SLOTS REPPED" in text
    return f"{len(rows)} reps, {len(done)}/{teams} slots, report renders"


@check("grade flags any club at the general stack threshold")
def _():
    """8/4: the IND cap covered Indianapolis and nothing else, so a mock with
    three Bears (all W10) passed clean. Any club reaching flag_at must be
    named in the grade, with its bye, and a named cap must not double-report."""
    from ffcli.draft import grade, general_cap, team_counts
    flag = general_cap()["flag_at"]
    stack = [{"round": i + 1, "pos": "WR", "team": "CHI", "player": f"Bear{i}"} for i in range(flag)]
    text = grade(stack, "LATE")
    assert "TEAM CONCENTRATION CHI" in text, f"a {flag}-Bear roster did not flag"
    assert f"all out W{__import__('ffcli.config', fromlist=['x']).bye_of('CHI')}" in text, \
        "concentration flag omits the shared bye week"
    assert "TEAM SPREAD" in text, "grade lost the spread summary"
    # one under the threshold stays quiet
    quiet = grade(stack[:-1], "LATE")
    assert "TEAM CONCENTRATION" not in quiet, f"{flag - 1} players should not flag"
    # a named cap reports once, as a cap - not also as a concentration
    colts = [{"round": i + 1, "pos": "WR", "team": "IND", "player": f"Colt{i}"} for i in range(flag)]
    ctext = grade(colts, "LATE")
    assert "BREACH" in ctext, "over-cap IND did not breach"
    assert "TEAM CONCENTRATION IND" not in ctext, "IND double-reported as cap AND concentration"
    assert team_counts(colts)["IND"] == flag
    return f"flags at {flag}, quiet at {flag - 1}, named caps report once"


@check("qb_board is coherent: six elite arms, real teams, windows match the rule")
def _():
    """The QB tiers say WHO for each window the qb_rule times. Six named elite
    arms (the branch map's premise), every team resolves to a real bye, the
    QB2/QB3 windows agree with qb_rule, and every never-list arm has a reason.
    Sheet must print the tiers so the table copy is never memory-dependent."""
    from ffcli.config import load, bye_of
    from ffcli.draft import sheet
    qb = load("qb_board")
    rule = load("qb_rule")
    assert len(qb["elite"]["who"]) == 6, f"elite tier has {len(qb['elite']['who'])} arms, needs 6"
    groups = [qb["elite"]["who"], qb["tier2_qb1"]["who"], qb["qb2_window"]["who"],
              qb["qb3_vets"]["who"], [qb["qb3_vets"]["fallback"]], qb["never"]]
    for p in (p for g in groups for p in g):
        assert bye_of(p["team"]), f"{p['player']}: team {p['team']!r} has no bye - not a real team code"
    assert qb["qb2_window"]["rounds"] == [rule["qb2_earliest_round"], rule["hard_floor_round"]], \
        "qb2_window rounds disagree with qb_rule"
    assert qb["qb3_vets"]["rounds"] == rule["qb3_rounds"], "qb3_vets rounds disagree with qb_rule"
    for p in qb["never"]:
        assert p.get("why"), f"never-list {p['player']} has no reason"
    text = sheet(1)
    for token in ("ELITE SIX", "QB2 ORDER", "QB3 VETS", "NEVER:"):
        assert token in text, f"sheet missing {token}"
    for p in qb["elite"]["who"]:
        assert p["player"] in text, f"sheet missing elite arm {p['player']}"
    return "6 elite arms, all teams resolve, windows match qb_rule, tiers on the sheet"


@check("two-column sheet fits one page and drops no rule")
def _():
    """The printable sheet chosen 8/4. Compression is where lessons die, so
    every hard rule, tally, tier and danger week must survive it - and it must
    stay inside one landscape page (119 cols, <=60 lines) for every branch."""
    import re
    from ffcli.draft import sheet_twocol, tree, picks_for_slot
    from ffcli.config import league, load, byes
    teams = league()["teams"]
    seen = set()
    for slot in range(1, teams + 1):
        label = tree(slot)["label"]
        if label in seen:
            continue
        seen.add(label)
        text = sheet_twocol(slot)
        lines = text.splitlines()
        assert len(lines) <= 60, f"{label}: {len(lines)} lines - past one page"
        wide = [l for l in lines if len(l) > 119]
        assert not wide, f"{label}: {len(wide)} lines wider than 119 cols"
        for token in ("HARD RULES", "ONE TE default", "QBs GONE", "BYES USED - CAP 2",
                      "QB1 BRANCH MAP", "QB TIERS", "NEVER", "STARS", "DANGER",
                      "FADE", "TIEBREAK", "TIER GAP", "GATE", "RUN:", "QB4", "K:"):
            assert token in text, f"{label}: two-col sheet lost {token}"
        # every bye week must show its teams, not just empty boxes
        for w, tms in byes().items():
            assert re.search(rf"W{w}\s*\[_\]\[_\] {re.escape(tms[0])}", text), \
                f"{label}: bye week W{w} missing its team list"
        # the QB tier names have to physically reach the page
        for p in load("qb_board")["elite"]["who"] + load("qb_board")["qb2_window"]["who"]:
            assert p["player"].split()[-1] in text, f"{label}: sheet lost {p['player']}"
        # slot-specific pick numbers, so the sheet is usable at the table
        assert str(picks_for_slot(slot)[0]) in text, f"{label}: no pick numbers"
    return f"{len(seen)} branches, <=60 lines, 119 cols, all rules + byes + tiers survive"


@check("snake pick numbers are right at both ends of the board")
def _():
    """picks_for_slot drives the sheet's pick column. Slot 1 and slot N are
    where an off-by-one shows up, and a wrong pick number at the table is
    worse than none."""
    from ffcli.draft import picks_for_slot
    from ffcli.config import league
    teams = league()["teams"]
    assert picks_for_slot(1)[:4] == [1, 24, 25, 48], picks_for_slot(1)[:4]
    assert picks_for_slot(teams)[:4] == [teams, teams + 1, 3 * teams, 3 * teams + 1], \
        picks_for_slot(teams)[:4]
    for slot in range(1, teams + 1):
        pk = picks_for_slot(slot)
        assert len(pk) == 17 and pk == sorted(pk), f"slot {slot}: picks not ascending"
        assert all(1 <= p <= 17 * teams for p in pk), f"slot {slot}: pick out of range"
    return f"slots 1-{teams} produce 17 ascending picks each"


@check("draft sheet is self-sufficient for a live draft")
def _():
    """Regression from the 7/25 mock: the sheet must work with no back-and-forth.
    Byes on every named target (a 4-player W14 stack slipped through without
    them), a tally grid, danger weeks, and the R1 pivot all printed."""
    import re
    from ffcli.draft import sheet, tree
    from ffcli.config import league
    seen = set()
    for slot in range(1, league()["teams"] + 1):
        label = tree(slot)["label"]
        if label in seen:
            continue
        seen.add(label)
        text = sheet(slot)
        for token in ("BYE TALLY", "QB COUNT TALLY", "DANGER WEEKS", "ROUND SCRIPT", "QB1 BRANCH MAP",
                      "RB FLOOR GATE", "WARREN GATE", "RUN TRIGGER", "FADES", "PERSONAL STARS",
                      "JOSH RULE", "TRIANGULATION"):
            assert token in text, f"{label}: sheet missing {token}"
        from ffcli.draft import _qb_trigger_rows
        from ffcli.config import load as _ld
        for _, gone, _a in _qb_trigger_rows(_ld("qb_rule")):
            assert f"{gone}[!]" in text, f"{label}: QB tally missing trigger mark at {gone}"
        bare = [ln for ln in text.splitlines()
                if "target:" in ln and not re.search(r"bye W\d+", ln)]
        assert not bare, f"{label}: target lines without a bye: {bare[:2]}"
        assert re.search(r"W14.*(ARI|DAL)", text), f"{label}: W14 playoff danger not spelled out"
    return f"{len(seen)} branches, byes on every target line"


@check("room model is structurally complete")
def _():
    """room.yaml must hold Josh + 11 managers with every profile field, the
    league reads, and slot_2026 fields ready for draft morning. TBD entries
    are allowed (the fill is human work) but counted so progress is visible."""
    from ffcli.config import load
    from ffcli.draft import room_report
    room = load("room")
    assert len(room["managers"]) == 11, f"expected 11 managers, found {len(room['managers'])}"
    for m in room["managers"]:
        for k in ("name", "qb_habit", "leans", "attention", "trades", "notes", "slot_2026"):
            assert k in m, f"manager {m.get('name')} missing '{k}'"
    for k in ("sharpest", "qb_run_2025", "patterns", "draft_order_method"):
        assert k in room["league"], f"league reads missing '{k}'"
    text = room_report(5)
    assert "SLOT 5 GEOMETRY" in text and "R5/R6 turn: picks 53 and 68" in text, \
        "slot geometry math wrong for slot 5 (DRAFT_BOARD_2025: slots 5-8 pick 53-56 then 65-68)"
    filled = sum(1 for m in room["managers"] if not str(m["name"]).startswith("TBD"))
    return f"{filled}/11 profiled, geometry checks out"


@check("2026 seating is a complete permutation and the seat intel renders")
def _():
    """The posted order assigns every manager a distinct seat 1-12 with Josh's
    slot matching league.yaml. A duplicate or missing seat silently corrupts
    every gap read in room_report, so it has to be a real permutation."""
    from ffcli.config import load
    from ffcli.draft import room_report
    room, lg = load("room"), load("league")
    mine = room["me"]["slot_2026"]
    assert mine == lg["draft"]["slot"], \
        f"room.me.slot_2026={mine} disagrees with league draft.slot={lg['draft']['slot']}"
    seats = [m["slot_2026"] for m in room["managers"]] + [mine]
    assert sorted(seats) == list(range(1, lg["teams"] + 1)), \
        f"2026 seats are not a permutation of 1-{lg['teams']}: {sorted(seats)}"
    intel = room["league"].get(f"slot{mine}_neighbors")
    assert isinstance(intel, list) and len(intel) >= 2, \
        f"no seat intel list for the drafted slot {mine}"
    text = room_report(mine)
    assert f"SEAT {mine} INTEL" in text, "seat intel block did not render in ff room"
    for para in intel:
        assert " ".join(str(para).split()) in text, "seat intel paragraph dropped from ff room"
    return f"seats 1-{lg['teams']} all distinct, Josh at {mine}, {len(intel)} intel notes render"


@check("target board covers 17 picks and matches the ledger and byes")
def _():
    """targets.yaml is the round-by-round WHO. It must stay consistent with
    the things it is derived from: one entry per round with the right pick
    number, a position skeleton that sums to the commitments ledger, and every
    named player's bye agreeing with byes.yaml. If a board name is retagged and
    this file is not, the draft-day list is silently wrong."""
    from ffcli.config import load, bye_of
    from ffcli.draft import targets_report, picks_for_slot
    t, lg = load("targets"), load("league")
    slot, rounds = lg["draft"]["slot"], lg["draft"]["rounds"]
    assert len(t["rounds"]) == rounds, f"expected {rounds} rounds, found {len(t['rounds'])}"
    expected = picks_for_slot(slot)
    for i, r in enumerate(t["rounds"]):
        assert r["rnd"] == i + 1, f"rounds out of order at index {i}"
        assert r["pick"] == expected[i], \
            f"R{r['rnd']} pick {r['pick']} != slot-{slot} pick {expected[i]}"
    # The need column has to spend exactly the ledger, no more and no less.
    ledger, spend = load("commitments")["ledger"], {}
    for r in t["rounds"]:
        assert r["pos"] in ledger, f"R{r['rnd']} has unknown pos {r['pos']!r}"
        spend[r["pos"]] = spend.get(r["pos"], 0) + 1
    assert spend == ledger, f"target board spends {spend}, ledger wants {ledger}"
    # The RB floor gates must be reachable on the skeleton as written.
    rb = [r["rnd"] for r in t["rounds"] if r["pos"] == "RB"]
    for by_round, need in ((4, 2), (8, 3), (12, 5)):
        held = sum(1 for x in rb if x <= by_round)
        assert held >= need, f"skeleton holds {held} RBs by R{by_round}, floor needs {need}"
    # Every named player's bye must agree with byes.yaml.
    for r in t["rounds"]:
        for p in list(r["take"]) + list(r.get("avoid", [])):
            if p["team"] == "-":
                continue
            real = bye_of(p["team"])
            assert real == p["bye"], \
                f"R{r['rnd']} {p['player']} ({p['team']}) tagged W{p['bye']}, byes.yaml says W{real}"
    text = targets_report()
    assert "CONFLICTS" in text and "R17 (pick 197)" in text, "targets_report did not render"
    named = sum(len(r["take"]) + len(r.get("avoid", [])) for r in t["rounds"])
    return f"{rounds} rounds, {named} named entries, ledger {spend}, byes agree"


@check("QB tier report catches an arm bought above its market band")
def _():
    """ff grade scores WHEN a QB was taken, never WHICH tier filled the slot.
    The 8/16 slot-5 rep scored 11/11 while spending pick 20 on a qb3_vets arm
    priced at picks 101-107. Every window hit, so nothing objected. This check
    guards the fix: a reach must be named a reach, and a correctly-priced room
    must stay silent."""
    from ffcli.draft import parse_picks, qb_tier_report, qb_tier_price
    reach = parse_picks("2 QB JAX Trevor Lawrence\n6 QB DET Jared Goff\n9 QB GB Jordan Love")
    text = "\n".join(qb_tier_report(reach))
    assert "REACH by 7" in text, f"7-round reach on a QB3-vet arm not flagged:\n{text}"
    assert "VALUE" in text, f"a QB2-tier arm taken at R9 should read as value:\n{text}"
    ok = parse_picks("1 QB BUF Josh Allen\n6 QB TEN Cam Ward\n9 QB NYJ Geno Smith")
    clean = "\n".join(qb_tier_report(ok))
    assert "REACH" not in clean and "above tier" not in clean, \
        f"correctly-priced QB room produced a false positive:\n{clean}"
    assert clean.count("at market") == 3, f"expected 3 at-market arms:\n{clean}"
    off = qb_tier_price("Bryce Young")
    assert off is None, f"an off-board arm resolved to a tier: {off}"
    assert qb_tier_price("Josh Allen")[0] == "ELITE SIX"
    return "reach, value, at-market and off-board all classified"


@check("live bye tally names the weeks at cap and the teams that blocks")
def _():
    """bye_stack is the oldest unfixed error - 13 of 18 reps - and it survives
    because the tally lives on the printed sheet as empty boxes nobody fills in
    mid-draft. The last four reps each broke a week with the LAST body added,
    three of them the kicker. This pins the on-screen tally: an over-cap week
    must be named, an at-cap week must block every team on that bye, and a
    clean W14 must say so."""
    from ffcli.draft import bye_block, draft_screen
    # The 8/16 rep's roster at the kicker pick: W5 and W6 already over cap.
    held = ["NE", "LV", "KC", "CAR", "LAC", "MIN", "DAL", "NE", "CAR",
            "MIN", "CHI", "HOU", "DET", "PIT", "NYJ"]
    text = "\n".join(bye_block(held))
    assert "W5" in text and "OVER CAP" in text, f"over-cap week not flagged:\n{text}"
    assert "W11  2" in text and "AT CAP" in text, f"at-cap week not flagged:\n{text}"
    blocked = next(l for l in text.splitlines() if "DO NOT DRAFT" in l)
    for t in ("MIN", "IND", "KC", "CAR"):
        assert t in blocked, f"{t} bye is at/over cap but it is not blocked:\n{blocked}"
    assert "SEEDING WEEK" in text, "a W14 body must be called out as the seeding week"
    # A clean roster: nothing at cap, so nothing is blocked, and W14 reads clear.
    clean = "\n".join(bye_block(["NE", "KC", "HOU"]))
    assert "DO NOT DRAFT" not in clean, f"false positive on a clean roster:\n{clean}"
    assert "W14 clear" in clean, f"clean W14 not confirmed:\n{clean}"
    assert bye_block([]) == [], "no teams should produce no block"
    # And it has to actually reach the pick screen.
    assert "BYES HELD" in draft_screen(5, 16, 24, teams=held), "tally missing from ff draft"
    assert "BYES HELD" not in draft_screen(5, 16, 24), "tally shown without --teams"
    return "over-cap, at-cap, blocked teams, seeding week and clean case all covered"


@check("K/DST board obeys the seeding week and the shared-bye rule")
def _():
    """These are two MANDATORY starters and the framework named zero of them
    until 8/16 - targets.yaml literally read 'Best available K'. The kicker
    created or worsened a bye breach in three of the last four reps, and the
    consensus K1 (Aubrey, DAL) is on the seeding week. This pins the rules that
    replaced 'best available': no W14, no W13 while Warren+Downs are planned,
    and the default K and DST must not share a bye."""
    from ffcli.config import load, bye_of
    from ffcli.draft import targets_report
    b = load("k_dst_board")
    for sec in ("kickers", "dst"):
        for grp in ("who", "refused"):
            for p in b[sec].get(grp, []):
                assert bye_of(p["team"]) == p["bye"], \
                    f"{p['player']} ({p['team']}) tagged W{p['bye']}, byes.yaml says W{bye_of(p['team'])}"
    # Nothing recommended may sit on the seeding week, or on W13 - Warren and
    # Downs are both IND and spend that week before R16.
    for sec in ("kickers", "dst"):
        for p in b[sec]["who"]:
            assert p["bye"] != 14, f"{p['player']} is recommended and on the SEEDING WEEK"
            assert p["bye"] != 13, f"{p['player']} is recommended and on W13, which Warren+Downs spend"
    # Aubrey is the consensus K1 and must be explicitly refused, not omitted.
    refused = {p["player"] for p in b["kickers"]["refused"]}
    assert "Brandon Aubrey" in refused, "the consensus K1 on the seeding week is not on the refuse list"
    # The default pairing must not share a bye week.
    k1 = b["kickers"]["who"][0]
    dst_default = next(d for d in b["dst"]["who"] if "DEFAULT" in str(d["verdict"]))
    assert k1["bye"] != dst_default["bye"], \
        f"default K {k1['player']} (W{k1['bye']}) and DST {dst_default['player']} share a bye"
    assert b["order"]["r16"] == "DST" and b["order"]["r17"] == "K", "R16/R17 order not recorded"
    # And the names have to reach the pick screen.
    for rnd, name in ((16, "Houston"), (17, "Cameron Dicker")):
        text = targets_report(rnd)
        assert name in text, f"R{rnd} target board does not name {name}"
        assert "Best available" not in text, f"R{rnd} still says 'best available'"
    return f"{len(b['kickers']['who'])} K + {len(b['dst']['who'])} DST named, byes verified"


@check("QB run trigger fires on rate and overrides count")
def _():
    """2025: the count sat at 8 for fourteen picks then hit 14 in twelve.
    Level-based thresholds cannot catch that; 3-in-12 rate can."""
    from ffcli.draft import qb_verdict
    hot = qb_verdict(5, 8, window=3)
    assert hot.action == "TAKE_QB2_NOW" and "RUN" in hot.note, \
        f"rate trigger did not fire: {hot.action}"
    cold = qb_verdict(5, 8, window=2)
    assert cold.action != "TAKE_QB2_NOW", f"trigger fired below threshold: {cold.action}"
    late = qb_verdict(8, 20, window=5)
    assert late.action == "PAST_FLOOR", "past-floor verdict must win over the rate trigger"
    return "fires at 3-in-12, quiet at 2, floor still wins"


@check("RB floor rule enforces 2/3/5 at rounds 4/8/12")
def _():
    """NEW 8/1: the 2025 failure (first RB at pick 65) had no rule to stop
    it. Mirror of the QB count rule."""
    from ffcli.draft import rb_verdict
    assert rb_verdict(4, 1).action == "TAKE_RB_NOW", "short at R4 gate not flagged"
    assert rb_verdict(4, 2).action == "ON_TRACK", "meeting the R4 gate flagged anyway"
    assert rb_verdict(8, 2).action == "TAKE_RB_NOW", "short at R8 gate not flagged"
    assert rb_verdict(12, 4).action == "PRIORITIZE_RB", "short at R12 gate not flagged"
    assert rb_verdict(12, 5).action == "ON_TRACK", "5 held at R12 flagged anyway"
    return "gates fire short, quiet when met"


@check("commitments encode the ledger: RB floor, QB2 window, K/DST last")
def _():
    """Structural guards from ROSTER_LEDGER.md: two RB commitments inside
    R1-4, QB2 no earlier than qb2_earliest_round, K and DST in R16-17."""
    from ffcli.config import load
    common = load("commitments")["common"]
    rbs_by_4 = [c for c in common if c["pick"].startswith("RB") and c["window"][1] <= 4]
    assert len(rbs_by_4) >= 2, f"RB floor not encoded: {len(rbs_by_4)} RB commitments by R4"
    qb2 = next(c for c in common if c["pick"].startswith("QB2"))
    earliest = load("qb_rule")["qb2_earliest_round"]
    assert qb2["window"][0] >= earliest, \
        f"QB2 window opens R{qb2['window'][0]}, earlier than qb2_earliest_round {earliest}"
    for pos in ("K", "DST"):
        c = next(c for c in common if c["pick"].split()[0] == pos)
        assert c["window"][0] >= 16, f"{pos} committed before R16"
    qb3 = next(c for c in common if c["pick"].startswith("QB3"))
    assert qb3["window"][0] >= 9 and qb3["window"][1] <= 13, \
        f"QB3 window {qb3['window']} outside the R9-13 ladder band (QB_LADDER.md)"
    assert not any(c["pick"].startswith("QB4") for c in common), \
        "QB4 must never be a draft commitment - in-season only (QB_LADDER.md)"
    return f"{len(rbs_by_4)} RBs by R4, QB2 opens R{qb2['window'][0]}, QB3 R{qb3['window'][0]}-{qb3['window'][1]}, K/DST R16+"


@check("grade scores a mock draft correctly")
def _():
    """A perfect MIDDLE script hits every commitment; removing Downs and adding
    a third Colt reports the miss and the stack-cap breach; --oneqb demotes QB
    commitments to observation instead of misses."""
    from ffcli.draft import grade, parse_picks
    perfect = parse_picks(
        "1 RB ATL Bijan Robinson\n2 QB TB Baker Mayfield\n3 RB DET Jahmyr Gibbs\n"
        "4 WR NYJ Garrett Wilson\n5 TE IND Tyler Warren\n6 QB TEN Cam Ward\n"
        "7 WR IND Josh Downs\n8 RB NE Rhamondre Stevenson\n10 QB NO Tyler Shough\n"
        "16 K JAX Cam Little\n17 DST TB Buccaneers\n")
    rep = grade(perfect, "MIDDLE")
    assert "11/11 commitments hit" in rep, f"perfect script not 11/11: {rep.splitlines()[:14]}"
    assert "MISS" not in rep and "BREACH" not in rep, "false negatives on a perfect script"
    assert "triangulation ok" in rep, "distinct QB byes (W10/W9/W8) not confirmed"

    flawed = parse_picks(
        "1 RB ATL Bijan Robinson\n2 QB TB Baker Mayfield\n3 RB DET Jahmyr Gibbs\n"
        "4 WR IND Alec Pierce\n5 TE IND Tyler Warren\n6 QB IND Daniel Jones\n"
        "7 RB NE Rhamondre Stevenson\n10 QB NYJ Geno Smith\n"
        "16 K JAX Cam Little\n17 DST TB Buccaneers\n")
    rep = grade(flawed, "MIDDLE")
    assert "MISS" in rep and "Downs" in rep, "missed Downs not reported"
    assert "BREACH" in rep, "3 Colts vs cap 2 not flagged"
    assert "10/11 commitments hit" in rep, "flawed score wrong"
    assert "4 players out" in rep, \
        "bye audit must count duplicate teams as separate players (3 IND + NYJ on W13)"
    assert "TRIANGULATION FAIL" in rep and "W13" in rep, \
        "Jones + Geno sharing the W13 bye must fail triangulation"

    rep = grade(parse_picks("1 RB ATL Bijan Robinson\n3 RB DET Jahmyr Gibbs\n"
                            "4 WR NYJ Garrett Wilson\n5 TE IND Tyler Warren\n"
                            "7 WR IND Josh Downs\n8 RB NE Rhamondre Stevenson\n"
                            "16 K JAX Cam Little\n17 DST TB Buccaneers\n"), "MIDDLE", oneqb=True)
    assert "OBS" in rep and "8/8 commitments hit" in rep, \
        "oneqb should demote the 3 QB commitments to OBS and grade the rest 8/8"
    return "perfect 11/11, flawed 10/11 + BREACH + triangulation fail, oneqb demotes QBs"


@check("QB rule urgency never dips between a trigger round and the floor")
def _():
    """Regression: rnd == hard_floor_round fell past every threshold band and
    returned NORMAL, so Round 6 gave no urgency at any QB count."""
    from ffcli.draft import qb_verdict
    from ffcli.config import load
    floor = load("qb_rule")["hard_floor_round"]
    for gone in (0, 10, 14, 16, 20, 24):
        v = qb_verdict(floor, gone).action
        assert v != "NORMAL", f"R{floor} gone={gone} returned NORMAL at the hard floor"
    hot = [r for r in range(1, floor + 2) if qb_verdict(r, 24).action == "NORMAL"]
    assert not hot, f"NORMAL at max QB count in rounds {hot}"
    return f"floor R{floor} triggers at every count"


@check("dashboard counts every status value present in the data")
def _():
    """Regression: Status has three values but the dashboard only counted two,
    hiding every Trending row from the At a Glance totals."""
    from ffcli.config import load
    import ffcli.workbook as wbmod, inspect
    used = {s for s in ("Unsettled", "Trending", "Resolved")
            if f'"{s}"' in inspect.getsource(wbmod._start_here)}
    present = {r.get("status") for r in load("watchlist") if r.get("status")}
    missing = present - used
    assert not missing, f"status values in data but not counted on dashboard: {sorted(missing)}"
    return f"{len(present)} status values, all counted"


@check("no board note is silently truncated by an unquoted comma")
def _():
    """Regression, found 8/16. In a YAML FLOW mapping - {player: X, why: ...} -
    an unquoted scalar ENDS at the first comma. Everything after it parses as a
    second key with a null value, so the note is silently cut in half and the
    printed board loses the reasoning. Twelve entries across targets.yaml and
    k_dst_board.yaml were already truncated this way and all 48 checks passed on
    them: Drake Maye's line lost 'ESPN has him SF QB2', the R4 second-TE refusal
    lost 'Not here, not now'. A key with a null value and a space in it is never
    legitimate in these files, which makes it a clean signature to assert on."""
    import glob, yaml
    orphans: list[str] = []

    def walk(node, path):
        if isinstance(node, dict):
            for k, v in node.items():
                if v is None and isinstance(k, str) and " " in k:
                    orphans.append(f"{path} -> {k!r}")
                walk(v, f"{path}.{k}")
        elif isinstance(node, list):
            for i, v in enumerate(node):
                walk(v, f"{path}[{i}]")

    files = sorted(glob.glob(str(ROOT / "data" / "*.yaml")))
    assert files, "no data files found to scan"
    for f in files:
        with open(f) as fh:
            walk(yaml.safe_load(fh), Path(f).name)
    assert not orphans, (
        f"{len(orphans)} truncated note(s) - quote the value: " + "; ".join(orphans[:4])
    )
    # Same family, different trigger: YAML 1.1 reads a bare NO as boolean False,
    # so New Orleans silently stops being a team. Found 8/18 on depth_board's
    # swap_team. Any key naming a team must hold a real team code.
    miscast: list[str] = []

    def teams(node, path):
        if isinstance(node, dict):
            for k, v in node.items():
                if isinstance(k, str) and "team" in k.lower() and isinstance(v, bool):
                    miscast.append(f"{path}.{k} -> {v!r}")
                teams(v, f"{path}.{k}")
        elif isinstance(node, list):
            for i, v in enumerate(node):
                teams(v, f"{path}[{i}]")

    for f in files:
        with open(f) as fh:
            teams(yaml.safe_load(fh), Path(f).name)
    assert not miscast, f"team field is not a team code (quote it): {'; '.join(miscast[:4])}"
    return f"{len(files)} data files scanned, no truncated notes, team codes intact"


@check("R11-R15 bye trap stays true to the boards it was derived from")
def _():
    """depth_board.bye_capacity documents a HARD trap: an on-plan R1-R10 caps
    W6/W7/W11/W13, which makes ten of the sixteen R11-R15 names illegal and
    leaves R14 with no legal option at all. Prose goes stale silently, so
    nothing here is trusted - every count, name and choke-point claim is
    RECOMPUTED from byes.yaml and targets.yaml and compared. If the Aug 25 pass
    reprices a board and the trap moves, this fails and forces a re-derivation
    instead of leaving a confident note that is quietly wrong."""
    from ffcli.config import load, bye_of
    import collections
    bc = load("depth_board")["bye_capacity"]
    rounds = {r["rnd"]: r for r in load("targets")["rounds"]}
    base = bc["baseline"]["picks"]
    assert len(base) == 10, f"baseline should cover R1-R10, found {len(base)}"

    def capped(picks):
        t = collections.Counter(bye_of(p["team"]) for p in picks)
        return {w for w, n in t.items() if n >= 2}

    def split(rnd, cap):
        takes = [p for p in rounds[rnd]["take"] if p["team"] != "-"]
        blocked = [p for p in takes if bye_of(p["team"]) in cap]
        return blocked, len(takes) - len(blocked)

    # the baseline must be real: byes agree, and it is legal on arrival at R11
    for p in base:
        assert bye_of(p["team"]) == p["bye"], \
            f"baseline R{p['rnd']} {p['player']} tagged W{p['bye']}, byes.yaml disagrees"
    cap0 = capped(base)
    assert max(collections.Counter(bye_of(p["team"]) for p in base).values()) <= 2, \
        "baseline itself breaks the 2-per-week cap - it cannot demonstrate anything"
    assert sorted(cap0) == sorted(bc["weeks_at_cap_by_r11"]), \
        f"weeks at cap recompute to {sorted(cap0)}, file says {bc['weeks_at_cap_by_r11']}"

    # counts, names and the zero-option rounds all recompute
    got_counts, got_names, dry = [], [], []
    for rnd in range(11, 16):
        blocked, n_legal = split(rnd, cap0)
        got_counts.append({"rnd": rnd, "blocked": len(blocked), "of": len(blocked) + n_legal})
        got_names += [{"rnd": rnd, "player": p["player"], "team": p["team"], "bye": p["bye"]}
                      for p in blocked]
        if n_legal == 0:
            dry.append(rnd)
    assert got_counts == bc["blocked"]["counts"], \
        f"blocked counts recompute to {got_counts}, file says {bc['blocked']['counts']}"
    assert got_names == bc["blocked"]["names"], \
        f"{len(got_names)} blocked names recomputed, file lists {len(bc['blocked']['names'])}"
    assert dry == bc["blocked"]["rounds_with_no_legal_option"], \
        f"rounds with no legal option recompute to {dry}, file says " \
        f"{bc['blocked']['rounds_with_no_legal_option']}"

    # every choke point must actually do what it claims, and no more
    for cp in bc["choke_points"]:
        rnd = cp["round"]
        assert any(p["rnd"] == rnd and p["team"] == cp["baseline_team"] for p in base), \
            f"choke point {cp['id']} says R{rnd} is {cp['baseline_team']}, baseline disagrees"
        swapped = [dict(p, team=cp["swap_team"]) if p["rnd"] == rnd else p for p in base]
        cap1 = capped(swapped)
        for wk in cp["frees_weeks"]:
            assert wk in cap0 and wk not in cap1, \
                f"{cp['id']} claims it frees W{wk}; recompute says otherwise"
        for r in cp["fixes_rounds"]:
            assert r in dry and split(r, cap1)[1] > 0, \
                f"{cp['id']} claims it fixes R{r}; recompute says otherwise"
        # an empty or partial fixes_rounds must be honest, not just unstated
        for r in dry:
            if r not in cp["fixes_rounds"]:
                assert split(r, cap1)[1] == 0, \
                    f"{cp['id']} silently fixes R{r} but does not claim it"
    return (f"{len(got_names)} names blocked at baseline, R{dry} dry, "
            f"{len(bc['choke_points'])} choke points verified")

@check("grade runs the RB floor gates over the finished roster")
def _():
    """Regression from mock rep 21 (8/18). That roster ended 6/6 RB and scored
    10/11, but the backs landed R2/R3/R8/R11/R13/R15 - four by R12 against a
    floor of five - and NOTHING in the grade said so. ledger_report counts final
    totals, commitments police individual picks, and neither watches the
    accumulation curve that rb_rule actually specifies. A late-but-complete
    backfield read as clean, which is the precise failure the floor exists to
    prevent. Both directions are asserted so the report cannot be made
    unconditional."""
    from ffcli.draft import grade, parse_picks
    from ffcli.config import load
    gates = load("rb_rule")["rb_floor"]
    late = ("1 QB NE Drake Maye\n2 RB KC Kenneth Walker III\n3 RB LAC Omarion Hampton\n"
            "4 WR CAR Tetairoa McMillan\n5 TE IND Tyler Warren\n6 QB MIN Kyler Murray\n"
            "7 WR GB Christian Watson\n8 RB PIT Rico Dowdle\n9 QB CAR Bryce Young\n"
            "10 WR LAC Quentin Johnston\n11 RB CHI Kyle Monangai\n12 WR PHI Makai Lemon\n"
            "13 RB ARI Tyler Allgeier\n14 WR NO Jordyn Tyson\n15 RB DET Isiah Pacheco\n"
            "16 DST LAC Chargers\n17 K PIT Chris Boswell\n")
    rep = grade(parse_picks(late), "MIDDLE")
    assert "ledger met" in rep, "fixture should still satisfy the 17-spot ledger"
    assert "RB FLOOR BREACH" in rep, \
        "6/6 RB arriving late must not read as clean - this is rep 21's exact miss"
    assert "R12: 4 held, floor is 5" in rep, f"breach not named precisely: {rep}"
    assert "R2, R3, R8, R11, R13, R15" in rep, "the accumulation curve must be printed"

    # moving one back inside the gate clears it, and every earlier gate still holds
    ontime = late.replace("12 WR PHI Makai Lemon", "12 RB PHI Makai Lemon")
    rep = grade(parse_picks(ontime), "MIDDLE")
    assert "RB FLOOR BREACH" not in rep and "met at every gate" in rep, \
        "a compliant curve must pass - the report cannot be unconditional"

    # an early gate breaks on its own, not only the last one
    early = late
    for a, b in (("2 RB KC", "2 WR KC"), ("3 RB LAC", "3 WR LAC"),
                 ("12 WR PHI", "12 RB PHI"), ("10 WR LAC", "10 RB LAC")):
        early = early.replace(a, b)
    rep = grade(parse_picks(early), "MIDDLE")
    first = gates[0]["by_end_of_round"]
    assert f"R{first}: 0 held" in rep, f"the first gate must fire on its own: {rep}"

    # a partial script is NOT a finished roster. The 11-pick commitments fixture
    # legitimately lists three backs, and inventing a breach there is exactly how
    # this report first broke the suite.
    partial = parse_picks("1 RB ATL Bijan Robinson\n3 RB DET Jahmyr Gibbs\n"
                          "8 RB NE Rhamondre Stevenson\n17 DST TB Buccaneers\n")
    assert "RB FLOOR" not in grade(partial, "MIDDLE"), \
        "floor gates must stay silent on a script shorter than a full draft"
    return f"{len(gates)} gates on finished rosters; breach, clean and partial covered"

@check("bye partition is arithmetically forced and matches the roster size")
def _():
    """The finding of 8/19: nine bye weeks x cap 2 = 18 slots, minus W14's
    banned second slot = 17 usable, against a 17-pick roster. SLACK IS ZERO, so
    the cap is an exact partition rather than a ceiling. Every number is
    recomputed from byes.yaml, bye_rule.yaml and league.yaml - if the bye map,
    the cap or the roster size ever changes, the documented partition is wrong
    and this fails instead of quietly lying. Also guards the direction of the
    finding: a week at ZERO is as broken as a week at three, so the targets must
    account for every pick."""
    from ffcli.config import load
    br, lg, by = load("bye_rule"), load("league"), load("byes")
    cap = br["cap"]["max_per_week"]
    sw = br["seeding_week"]
    picks = lg["draft"]["rounds"]
    assert sw["week"] in by, f"seeding week W{sw['week']} is not a real bye week"
    usable = cap * len(by) - (cap - sw["max_players"])
    assert usable == picks, (
        f"partition broken: {len(by)} weeks x cap {cap} minus "
        f"{cap - sw['max_players']} banned seeding slot(s) = {usable}, "
        f"roster is {picks} picks")
    tgt = {t["week"]: t["exactly"] for t in br["partition"]["targets"]}
    assert set(tgt) == set(by), \
        f"targets cover {sorted(tgt)}, byes.yaml has {sorted(by)}"
    assert sum(tgt.values()) == picks, \
        f"targets sum to {sum(tgt.values())}, roster is {picks} - every pick must land somewhere"
    assert tgt[sw["week"]] == sw["max_players"], "seeding-week target contradicts the seeding rule"
    for w, n in tgt.items():
        assert 0 < n <= cap, f"W{w} target {n} is not between 1 and the cap"
    return f"{len(by)} weeks x {cap} - 1 = {usable} slots = {picks} picks, slack 0"


@check("board supply gaps match what the boards can actually field")
def _():
    """bye_rule.supply_gap says the named board cannot field the partition -
    W5 needs two bodies and offers ZERO that are not a fade or a demoted arm,
    which is why bye_stack_w5 is in five of six slot-5 reps and why Brooks has
    been drafted seven times. Prose about a gap goes stale the moment a board is
    repriced, so the supply is recomputed from targets.yaml the way a drafter
    reads it - take entries only, less fades and the never list - and compared
    against every short and thin week the file claims. Adding a real W5 name
    closes the gap and fails this until the file is updated."""
    from ffcli.config import load
    import collections
    br, T = load("bye_rule"), load("targets")
    fades = {f["player"] for f in load("depth_board").get("fades", [])}
    never = {e if isinstance(e, str) else e.get("player")
             for e in load("qb_board").get("never", [])}
    assert fades, "no fades loaded - the filter would be vacuous"
    supply = collections.defaultdict(set)
    for r in T["rounds"]:
        for p in r["take"]:
            if p["team"] == "-" or p["player"] in fades or p["player"] in never:
                continue
            supply[p["bye"]].add(p["player"])
    tgt = {t["week"]: t["exactly"] for t in br["partition"]["targets"]}
    claimed = {}
    for key in ("short", "thin"):
        for row in br["supply_gap"].get(key) or []:
            w = row["week"]
            claimed[w] = key
            assert row["need"] == tgt[w], \
                f"W{w} supply_gap says need {row['need']}, partition says {tgt[w]}"
            got = len(supply.get(w, ()))
            assert got == row["named_draftable"], \
                f"W{w} recomputes to {got} draftable names, file says {row['named_draftable']}"
            if key == "short":
                assert got < row["need"], f"W{w} listed short but supply {got} meets need {row['need']}"
            else:
                assert row["need"] <= got <= row["need"] + 1, \
                    f"W{w} listed thin but supply is {got} against need {row['need']}"
    # nothing genuinely short may go unlisted
    for w, need in tgt.items():
        if len(supply.get(w, ())) < need:
            assert claimed.get(w) == "short", f"W{w} is short and not listed in supply_gap"
    return (f"{len(claimed)} weeks tracked, "
            f"W5 supply {len(supply.get(5, ()))} vs need {tgt[5]}")

@check("grade scores the finished roster against the bye partition")
def _():
    """Rep 23 (8/19) ran four weeks OVER cap and four weeks UNDER, balanced
    exactly - one redistribution error, not four - and skipping W14 entirely is
    what forced the first breach, since eight weeks at cap 2 hold only 16 of its
    17 players. audit() reported the four overs and was silent on the four empty
    slots and on the cause, because it was written when the cap read as a
    ceiling. bye_rule.partition proves it is not one. Both directions are
    asserted so the report cannot become unconditional, and the exact case is
    pinned to data/ideal_roster.txt so a board reprice that breaks the ideal
    line fails here too."""
    from ffcli.draft import grade, parse_picks
    from ffcli.config import load
    from pathlib import Path
    tgt = {t["week"]: t["exactly"] for t in load("bye_rule")["partition"]["targets"]}
    seed = load("bye_rule")["seeding_week"]["week"]

    ideal = parse_picks(Path(ROOT / "data" / "ideal_roster.txt").read_text())
    assert len(ideal) == load("league")["draft"]["rounds"], "ideal roster is not a full draft"
    rep = grade(ideal, "MIDDLE")
    assert "BYE PARTITION: exact" in rep, \
        f"the solved ideal roster must hit the partition exactly: {rep}"

    # rep 23: four over, four under, W14 empty
    r23 = parse_picks(
        "1 RB DET Jahmyr Gibbs\n2 QB NYG Jaxson Dart\n3 RB KC Kenneth Walker III\n"
        "4 WR NYG Malik Nabers\n5 TE IND Tyler Warren\n6 QB TB Baker Mayfield\n"
        "7 WR TEN Carnell Tate\n8 RB NE Rhamondre Stevenson\n9 QB GB Jordan Love\n"
        "10 WR IND Josh Downs\n11 RB MIN Jordan Mason\n12 RB SEA Zach Charbonnet\n"
        "13 WR TB Jalen McMillan\n14 RB SF Kaelon Black\n15 WR IND Keenan Allen\n"
        "16 DST LAC Chargers\n17 K TB Chase McLaughlin\n")
    rep = grade(r23, "MIDDLE")
    assert "BYE PARTITION OFF by 4" in rep, f"rep 23 is off by 4: {rep}"
    assert "4 week(s) over, 4 week(s) under" in rep, "both directions must be counted"
    assert f"W{seed}  0 held, target {tgt[seed]}" in rep, \
        "an EMPTY seeding week must be reported - it is what forces the overflow"
    assert "SAME error" in rep, "over and under must be tied together, not listed apart"
    return f"exact and off-by-4 cases both covered, {len(tgt)} weeks targeted"

@check("K/DST are flagged when they land off the partition")
def _():
    """Rule r16_r17_are_partition_fillers, added 8/21 on four consecutive reps
    where one of the last two picks tipped a week over cap: McPherson made W6 a
    four, McLaughlin made W10 a three, the Vikings D/ST made W6 a three, the
    Chargers D/ST made W7 a three while W8 sat at ZERO. At R16 fifteen players
    are held, so the short weeks are fully determined and these two fungible
    picks should fill them. The grade must name it when they do not, and must
    stay quiet when they do - a nag that always fires would be ignored by R17."""
    from ffcli.draft import grade, parse_picks
    from ffcli.config import load
    from pathlib import Path
    rules = {r["id"]: r for r in load("k_dst_board")["rules"]}
    new = rules["r16_r17_are_partition_fillers"]
    assert new["severity"] == "HARD", "the partition-filler rule must be HARD"
    old = rules["bye_decides_among_equals"]
    assert str(old["severity"]).startswith("SUPERSEDED"), \
        "the weaker avoid-a-capped-week rule must be marked superseded, not left live"
    assert old["superseded_by"] == "r16_r17_are_partition_fillers", \
        "a superseded rule must name its replacement - never silently dropped"

    # rep 25: the DST sat on an over-target week while three weeks were short
    r25 = parse_picks(
        "1 RB DET Jahmyr Gibbs\n2 QB JAX Trevor Lawrence\n3 RB KC Kenneth Walker III\n"
        "4 WR BUF DJ Moore\n5 TE IND Tyler Warren\n6 QB MIN Kyler Murray\n"
        "7 WR IND Josh Downs\n8 RB TB Kenny Gainwell\n9 QB GB Jordan Love\n"
        "10 RB LAR Blake Corum\n11 RB CHI Kyle Monangai\n12 WR PHI Makai Lemon\n"
        "13 RB ARI Tyler Allgeier\n14 WR MIA Malik Washington\n15 WR MIN Jauan Jennings\n"
        "16 DST LAC Chargers\n17 K PIT Chris Boswell\n")
    rep = grade(r25, "MIDDLE")
    assert "partition fillers" in rep, f"the misplaced D/ST must be named: {rep}"
    assert "R16 DST" in rep and "W7" in rep, "the offending pick and its week must be identified"

    # The negative case must exercise the CONDITION, not the exact-partition
    # early return. Same rep 25 roster, still off the partition, but with the
    # defense moved onto a week that is short (HOU/W8) instead of one already
    # over (LAC/W7). The flag must go quiet even though the roster is still
    # wrong overall - an always-on nag would be ignored by R17.
    fixed = parse_picks("\n".join(
        ln.replace("16 DST LAC Chargers", "16 DST HOU Houston")
        for ln in """1 RB DET Jahmyr Gibbs
2 QB JAX Trevor Lawrence
3 RB KC Kenneth Walker III
4 WR BUF DJ Moore
5 TE IND Tyler Warren
6 QB MIN Kyler Murray
7 WR IND Josh Downs
8 RB TB Kenny Gainwell
9 QB GB Jordan Love
10 RB LAR Blake Corum
11 RB CHI Kyle Monangai
12 WR PHI Makai Lemon
13 RB ARI Tyler Allgeier
14 WR MIA Malik Washington
15 WR MIN Jauan Jennings
16 DST LAC Chargers
17 K PIT Chris Boswell""".splitlines()))
    rep = grade(fixed, "MIDDLE")
    assert "BYE PARTITION OFF" in rep, "fixture must still be off the partition overall"
    assert "partition fillers" not in rep, \
        "a K and D/ST on short weeks must not be nagged even on an imperfect roster"

    # and the solved ideal roster stays silent too
    ideal = parse_picks(Path(ROOT / "data" / "ideal_roster.txt").read_text())
    assert "partition fillers" not in grade(ideal, "MIDDLE"), "exact roster must be silent"
    return "misplaced K/DST named, correct placement silent, superseded rule retained"

@check("the R14-R15 W6 trap names real bodies on a real week")
def _():
    """Named 8/21 after reps 25 and 26 took the IDENTICAL two receivers at
    R14/R15 with the rounds swapped - Jennings (MIN) and Washington (MIA), both
    W6 - and rep 26 finished with five players on W6, the worst single-week hole
    ever recorded at this slot. The trap is only worth stating if every claim in
    it is still true, so the pair's byes, the week's crowding and the refusal
    itself are all recomputed rather than trusted."""
    from ffcli.config import load, bye_of
    import collections
    trap = load("depth_board")["w6_late_trap"]
    tgt = {t["week"]: t["exactly"] for t in load("bye_rule")["partition"]["targets"]}
    wk = trap["weeks"][0]
    assert trap["severity"] == "HARD", "taking both must be a hard refuse"

    # every named body must be real and must actually sit on the trap's week
    names = trap["names"]
    assert len(names) >= 2, "a trap about one player is just a fade"
    for p in names:
        real = bye_of(p["team"])
        assert real == p["bye"] == wk, \
            f"{p['player']} ({p['team']}) tagged W{p['bye']}, byes.yaml says W{real}, trap is W{wk}"
    assert len({p["team"] for p in names}) > 1, \
        "all from one club would be a stack-cap issue, not a bye trap"
    assert trap["rounds"] == sorted(trap["rounds"]) and all(
        1 <= r <= load("league")["draft"]["rounds"] for r in trap["rounds"]), \
        "the trap must name real, ordered rounds"

    # the week must genuinely be crowded on the boards, or the trap is folklore
    supply = collections.Counter()
    for r in load("targets")["rounds"]:
        for e in r["take"]:
            if e["team"] != "-":
                supply[e["bye"]] += 1
    assert supply[wk] > tgt[wk] * 2, \
        f"W{wk} offers only {supply[wk]} named bodies against a target of {tgt[wk]} - not a crowded week"
    busiest = max(supply, key=lambda w: supply[w])
    assert supply[wk] >= supply[busiest] * 0.6, \
        f"W{wk} is not among the crowded weeks (W{busiest} has {supply[busiest]})"
    return f"{len(names)} names verified on W{wk}, board offers {supply[wk]} vs target {tgt[wk]}"

@check("ADP pass measured-availability numbers recompute from the stored rosters")
def _():
    """The 8/25 pass is the first to carry MEASURED numbers rather than
    estimates, and it drove three repricings - Love to a QB3 price, Allgeier to
    R13, Boston onto the board. Those claims were unreproducible when written:
    mocks.yaml stores scores and notes but never stored the picks. The rosters
    now live in data/rep_rosters and every latest_pick in the pass is recomputed
    from them here, so a claim cannot outlive the evidence for it. Also asserts
    the rosters themselves are well formed, since a silently truncated roster
    would quietly weaken every number above."""
    from ffcli.config import load
    from ffcli.draft import parse_picks, picks_for_slot
    import glob
    rows = sorted(glob.glob(str(ROOT / "data" / "rep_rosters" / "*_slot5_*.txt")))
    ap = load("adp")["pass_2026_08_25"]
    n_claimed = ap["measured_availability"]["n_reps"]
    assert len(rows) == n_claimed, \
        f"pass claims {n_claimed} reps, data/rep_rosters holds {len(rows)}"

    rounds_per_draft = load("league")["draft"]["rounds"]
    pick_of = dict(enumerate(picks_for_slot(5), start=1))
    seen: dict[str, list[int]] = {}
    for f in rows:
        picks = parse_picks(Path(f).read_text())
        assert len(picks) == rounds_per_draft, \
            f"{Path(f).name} holds {len(picks)} picks, a draft is {rounds_per_draft}"
        assert [p["round"] for p in picks] == list(range(1, rounds_per_draft + 1)), \
            f"{Path(f).name} rounds are not 1..{rounds_per_draft} in order"
        for p in picks:
            seen.setdefault(p["player"], []).append(p["round"])

    for row in ap["measured_availability"]["who"]:
        name, claimed_pick, claimed_n = row["player"], row["latest_pick"], row["n"]
        got = seen.get(name)
        assert got, f"{name} is claimed in the pass but appears in no stored roster"
        assert len(got) == claimed_n, \
            f"{name}: pass says taken {claimed_n}x, rosters show {len(got)}"
        real = pick_of[max(got)]
        assert real == claimed_pick, \
            f"{name}: pass says last seen at pick {claimed_pick}, rosters say {real}"
    return f"{len(rows)} rosters, {len(ap['measured_availability']['who'])} claims recomputed"

# --------------------------------------------------------------- report
def report() -> int:
    width = max(len(n) for n, _, _ in results) + 2
    passed = sum(1 for _, ok, _ in results if ok)
    print("=" * (width + 40))
    print("ff2026 verification")
    print("=" * (width + 40))
    for name, ok, detail in results:
        mark = "PASS" if ok else "FAIL"
        print(f"[{mark}] {name:<{width}} {detail}")
    print("-" * (width + 40))
    print(f"{passed}/{len(results)} checks passed")
    if passed != len(results):
        print("\nStructural failures above. Fix before manual review.")
        return 1
    print("\nStructure is sound. This does NOT validate the football analysis -")
    print("see HANDOFF.md for what still needs human verification.")
    return 0


if __name__ == "__main__":
    sys.exit(report())
